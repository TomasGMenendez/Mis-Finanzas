"""
Capa de escritura al Excel para el CRM.

Todo lo que el CRM guarda pasa por acá. Es el único módulo que toca
Finanzas_Toto.xlsx desde el CRM, así que acá viven las protecciones:

  - Nunca escribe si el Excel está abierto en Excel.exe (se perderían los
    cambios de uno de los dos lados).
  - Antes de cada escritura hace una copia de seguridad en Programas/backups/.
  - Las posiciones de las secciones NO están hardcodeadas: se buscan por el
    texto de su encabezado, igual que hace generar_datos_html.py. Así, si en el
    futuro se agregan o mueven filas, esto sigue funcionando.
"""
import re
import shutil
from contextlib import contextmanager
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font

BASE = Path(__file__).parent
RAIZ = BASE.parent
XLSX = RAIZ / "Finanzas_Toto.xlsx"
BACKUPS = BASE / "backups"
MAX_BACKUPS = 40

# Estilo de la planilla (mismos valores que usa mercadopago_processor.py)
FN = "Segoe UI"
INK = "F5F5F5"
BLUE = "7CA8FF"

# Límite de filas de Gastos/Ingresos — es el mismo que usan las fórmulas de
# TOTAL de esas hojas (=SUM(J5:J504)). Pasarse rompería esos totales.
MAX_FILA_MOV = 504
PRIMERA_FILA_MOV = 5


class ExcelAbierto(Exception):
    """El Excel está abierto en Excel.exe y no se puede escribir."""


class ErrorCRM(Exception):
    """Cualquier error previsible que hay que mostrarle al usuario tal cual."""


# ---------------------------------------------------------------------------
# Seguridad: bloqueo y backups
# ---------------------------------------------------------------------------
def verificar_excel_cerrado():
    """Excel deja un archivo '~$nombre.xlsx' mientras tiene el libro abierto.

    Si escribiéramos igual, openpyxl reescribe el archivo entero y cuando el
    usuario después guarda desde Excel pisa todo lo que cargó el CRM (o al
    revés). Es preferible frenar y pedirle que lo cierre.
    """
    if (RAIZ / f"~${XLSX.name}").exists():
        raise ExcelAbierto(
            "El Excel está abierto. Guardalo y cerralo, y volvé a intentar."
        )


def hacer_backup():
    """Copia el Excel a Programas/backups/ antes de modificarlo.

    El CRM escribe sobre datos financieros reales, así que cada guardado deja
    una copia con fecha y hora. Se conservan las últimas MAX_BACKUPS.
    """
    BACKUPS.mkdir(exist_ok=True)
    sello = datetime.now().strftime("%Y%m%d_%H%M%S")
    destino = BACKUPS / f"Finanzas_Toto_{sello}.xlsx"
    shutil.copyfile(XLSX, destino)

    copias = sorted(BACKUPS.glob("Finanzas_Toto_*.xlsx"))
    for vieja in copias[:-MAX_BACKUPS]:
        vieja.unlink(missing_ok=True)
    return destino


@contextmanager
def editar():
    """Abre el libro, deja modificarlo y lo guarda con backup previo."""
    verificar_excel_cerrado()
    hacer_backup()
    wb = load_workbook(XLSX)
    yield wb
    # openpyxl guarda las fórmulas pero no sus resultados. Con esto, Excel
    # recalcula todo el libro la próxima vez que lo abra y no muestra celdas
    # vacías donde hay fórmulas.
    wb.calculation.fullCalcOnLoad = True
    try:
        wb.save(XLSX)
    except PermissionError:
        raise ExcelAbierto(
            "No pude guardar: el Excel está abierto o OneDrive lo está "
            "sincronizando. Cerralo y volvé a intentar."
        )


# ---------------------------------------------------------------------------
# Utilidades
# ---------------------------------------------------------------------------
def _fecha(valor):
    """Acepta 'YYYY-MM-DD' o un date/datetime y devuelve siempre un date."""
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    if isinstance(valor, str) and valor.strip():
        return datetime.strptime(valor.strip()[:10], "%Y-%m-%d").date()
    raise ErrorCRM("Falta la fecha.")


def _numero(valor, campo="monto"):
    try:
        n = float(str(valor).replace(",", "."))
    except (TypeError, ValueError):
        raise ErrorCRM(f"El {campo} tiene que ser un número.")
    if n == 0:
        raise ErrorCRM(f"El {campo} no puede ser cero.")
    return n


def _texto(valor):
    return str(valor).strip() if valor is not None else ""


def _orden_fecha(valor):
    """Clave de ordenamiento tolerante a fechas mezcladas con texto."""
    if isinstance(valor, datetime):
        return valor
    if isinstance(valor, date):
        return datetime.combine(valor, datetime.min.time())
    return datetime.max


def _rango_nombrado(wb, nombre):
    """Devuelve la lista de valores de un rango con nombre del Excel."""
    definicion = wb.defined_names.get(nombre)
    if definicion is None:
        return []
    m = re.match(r"'?([^'!]+)'?!\$?([A-Z]+)\$?(\d+):\$?([A-Z]+)\$?(\d+)",
                 definicion.attr_text)
    if not m:
        return []
    hoja, col1, fila1, _col2, fila2 = m.groups()
    ws = wb[hoja]
    valores = []
    for fila in range(int(fila1), int(fila2) + 1):
        v = ws[f"{col1}{fila}"].value
        if v is not None and str(v).strip():
            valores.append(str(v).strip())
    return valores


# ---------------------------------------------------------------------------
# Localización de las secciones de la hoja Inversiones
#
# Se buscan por el texto de su encabezado en vez de por número de fila fijo,
# para que agregar o mover filas en el Excel no rompa el CRM.
# ---------------------------------------------------------------------------
def _buscar_fila(ws, patron, col=2, desde=1):
    rx = re.compile(patron, re.I)
    for r in range(desde, ws.max_row + 1):
        v = ws.cell(r, col).value
        if v is not None and rx.match(str(v).strip()):
            return r
    return None


def localizar_secciones(ws):
    """Devuelve los rangos de cada bloque editable de la hoja Inversiones."""
    sec = {}

    # Bull Market: arranca después del divisor y termina en el subtotal, que
    # es la fila que tiene "BULL MARKET ARS →" en la columna G.
    div_bm = _buscar_fila(ws, r"──\s*BULL MARKET")
    if div_bm:
        # El bloque no tiene encabezados propios: reusa los de la fila 4 y los
        # datos arrancan en la fila siguiente al divisor.
        primera = div_bm + 1
        subtotal = None
        for r in range(primera, ws.max_row + 1):
            if str(ws.cell(r, 7).value or "").strip().upper().startswith("BULL MARKET"):
                subtotal = r
                break
        if subtotal:
            sec["bull_market"] = (primera, subtotal - 1)

    # Cripto: filas entre el encabezado (fila con "Activo") y el divisor de
    # Bull Market. Bybit las sincroniza solo, pero el CRM las muestra.
    hdr = _buscar_fila(ws, r"^Activo$")
    if hdr and div_bm:
        sec["cripto"] = (hdr + 1, div_bm - 1)

    # Aportes de inversión
    div_ap = _buscar_fila(ws, r"──\s*APORTES")
    if div_ap:
        hdr_ap = None
        for r in range(div_ap, min(div_ap + 10, ws.max_row + 1)):
            if _texto(ws.cell(r, 2).value) == "Fecha" and _texto(ws.cell(r, 4).value) == "Monto":
                hdr_ap = r
                break
        if hdr_ap:
            fin = hdr_ap + 1
            while fin <= ws.max_row and not str(ws.cell(fin, 6).value or "").upper().startswith("TOTAL"):
                fin += 1
            sec["aportes"] = (hdr_ap + 1, fin - 1)

    # Registro de ventas de mercadería
    div_v = _buscar_fila(ws, r"──\s*REGISTRO DE VENTAS")
    if div_v:
        hdr_v = _buscar_fila(ws, r"^Fecha venta$", desde=div_v)
        if hdr_v:
            fin = hdr_v + 1
            while fin <= ws.max_row and not str(ws.cell(fin, 2).value or "").upper().startswith("TOTAL"):
                fin += 1
            sec["ventas"] = (hdr_v + 1, fin - 1)

    return sec


def _primera_libre(ws, desde, hasta, col=2):
    for r in range(desde, hasta + 1):
        if ws.cell(r, col).value is None:
            return r
    return None


# ---------------------------------------------------------------------------
# Lectura: catálogos y estado actual (para poblar los formularios del CRM)
# ---------------------------------------------------------------------------
def leer_catalogos():
    wb = load_workbook(XLSX, data_only=True)
    ing = wb["Ingresos"]
    fuentes = []
    for r in range(PRIMERA_FILA_MOV, MAX_FILA_MOV + 1):
        v = ing.cell(r, 4).value
        if v and str(v).strip() not in fuentes:
            fuentes.append(str(v).strip())

    metas = []
    hm = wb["Metas"]
    for r in range(5, hm.max_row + 1):
        v = hm.cell(r, 2).value
        if v and str(v).strip():
            metas.append(str(v).strip())

    datos = {
        "categorias": _rango_nombrado(wb, "CATEGORIAS"),
        "medios": _rango_nombrado(wb, "MEDIOS"),
        "tipos": _rango_nombrado(wb, "TIPOS"),
        "monedas": _rango_nombrado(wb, "MONEDAS"),
        "tiposInversion": _rango_nombrado(wb, "TIPOS_INV"),
        "fuentes": sorted(fuentes),
        "metas": metas,
        "usdtArs": wb["Config"]["C5"].value or 0,
        "dolarBlue": wb["Config"]["C6"].value or 0,
    }
    wb.close()
    return datos


def leer_estado():
    """Resumen liviano para mostrar en el CRM: últimos movimientos y posiciones."""
    wb = load_workbook(XLSX, data_only=True)
    gas, ing, inv = wb["Gastos"], wb["Ingresos"], wb["Inversiones"]

    def _iso(v):
        return v.strftime("%Y-%m-%d") if isinstance(v, (date, datetime)) else None

    gastos = []
    for r in range(PRIMERA_FILA_MOV, MAX_FILA_MOV + 1):
        f = _iso(gas.cell(r, 2).value)
        if f and gas.cell(r, 9).value:
            gastos.append({
                "fila": r, "fecha": f,
                "categoria": _texto(gas.cell(r, 4).value),
                "descripcion": _texto(gas.cell(r, 5).value),
                "medio": _texto(gas.cell(r, 6).value),
                "moneda": _texto(gas.cell(r, 8).value) or "ARS",
                "monto": gas.cell(r, 9).value,
            })

    ingresos = []
    for r in range(PRIMERA_FILA_MOV, MAX_FILA_MOV + 1):
        f = _iso(ing.cell(r, 2).value)
        if f and ing.cell(r, 7).value:
            ingresos.append({
                "fila": r, "fecha": f,
                "fuente": _texto(ing.cell(r, 4).value),
                "descripcion": _texto(ing.cell(r, 5).value),
                "moneda": _texto(ing.cell(r, 6).value) or "ARS",
                "monto": ing.cell(r, 7).value,
            })

    sec = localizar_secciones(inv)

    posiciones = []
    if "bull_market" in sec:
        ini, fin = sec["bull_market"]
        for r in range(ini, fin + 1):
            if inv.cell(r, 2).value:
                posiciones.append({
                    "fila": r,
                    "activo": _texto(inv.cell(r, 2).value),
                    "tipo": _texto(inv.cell(r, 3).value),
                    "fechaCompra": _iso(inv.cell(r, 4).value),
                    "cantidad": inv.cell(r, 5).value,
                    "precioCompra": inv.cell(r, 6).value,
                    "precioActual": inv.cell(r, 7).value,
                    "meta": _texto(inv.cell(r, 12).value),
                })

    aportes = []
    if "aportes" in sec:
        ini, fin = sec["aportes"]
        for r in range(ini, fin + 1):
            f = _iso(inv.cell(r, 2).value)
            if f and inv.cell(r, 4).value:
                aportes.append({
                    "fila": r, "fecha": f,
                    "monto": inv.cell(r, 4).value,
                    "moneda": _texto(inv.cell(r, 5).value) or "ARS",
                    "nota": _texto(inv.cell(r, 6).value),
                })

    ventas = []
    if "ventas" in sec:
        ini, fin = sec["ventas"]
        for r in range(ini, fin + 1):
            f = _iso(inv.cell(r, 2).value)
            if f and inv.cell(r, 4).value:
                ventas.append({
                    "fila": r, "fecha": f,
                    "producto": _texto(inv.cell(r, 3).value),
                    "cantidad": inv.cell(r, 4).value,
                    "precioVenta": inv.cell(r, 5).value,
                })

    libres = 0
    if "bull_market" in sec:
        ini, fin = sec["bull_market"]
        libres = sum(1 for r in range(ini, fin + 1) if not inv.cell(r, 2).value)

    wb.close()
    gastos.sort(key=lambda g: g["fecha"], reverse=True)
    ingresos.sort(key=lambda i: i["fecha"], reverse=True)
    aportes.sort(key=lambda a: a["fecha"], reverse=True)
    ventas.sort(key=lambda v: v["fecha"], reverse=True)
    return {
        "gastos": gastos, "ingresos": ingresos, "posiciones": posiciones,
        "aportes": aportes, "ventas": ventas, "lugaresLibresBM": libres,
    }


# ---------------------------------------------------------------------------
# Escritura: Gastos e Ingresos
#
# Las dos hojas se mantienen siempre en orden cronológico (más vieja arriba),
# que es como Toto pidió que queden. Por eso, después de agregar, se reescribe
# la hoja entera ordenada por fecha y se regeneran las fórmulas de cada fila.
# ---------------------------------------------------------------------------
_CAMPOS_GASTO = ("fecha", "categoria", "descripcion", "medio", "tipo", "moneda", "monto")
_CAMPOS_INGRESO = ("fecha", "fuente", "descripcion", "moneda", "monto")


def _leer_filas(ws, columnas):
    filas = []
    for r in range(PRIMERA_FILA_MOV, MAX_FILA_MOV + 1):
        if ws.cell(r, 2).value is None:
            continue
        filas.append({nombre: ws.cell(r, col).value for nombre, col in columnas.items()})
    return filas


def _escribir_gastos(ws, filas):
    filas.sort(key=lambda f: _orden_fecha(f["fecha"]))
    if len(filas) > MAX_FILA_MOV - PRIMERA_FILA_MOV + 1:
        raise ErrorCRM("La hoja Gastos se quedó sin filas libres.")
    for r in range(PRIMERA_FILA_MOV, MAX_FILA_MOV + 1):
        for c in range(2, 12):
            ws.cell(r, c).value = None
    for i, f in enumerate(filas):
        r = PRIMERA_FILA_MOV + i
        ws.cell(r, 2, f["fecha"]).number_format = "dd/mm/yyyy"
        ws.cell(r, 3, f'=IF(B{r}="","",TEXT(B{r},"mmm-yy"))')
        for col, clave in ((4, "categoria"), (5, "descripcion"), (6, "medio"),
                           (7, "tipo"), (8, "moneda")):
            ws.cell(r, col, f[clave]).font = Font(name=FN, color=INK, size=10)
        ws.cell(r, 9, f["monto"]).font = Font(name=FN, color=BLUE)
        ws.cell(r, 10, f'=IF(OR(I{r}="",H{r}=""),"",IF(H{r}="ARS",I{r},I{r}*USDT_ARS))')
        ws.cell(r, 11, f'=IF(OR(I{r}="",H{r}=""),"",IF(H{r}="USDT",I{r},I{r}/USDT_ARS))')


def _escribir_ingresos(ws, filas):
    filas.sort(key=lambda f: _orden_fecha(f["fecha"]))
    if len(filas) > MAX_FILA_MOV - PRIMERA_FILA_MOV + 1:
        raise ErrorCRM("La hoja Ingresos se quedó sin filas libres.")
    for r in range(PRIMERA_FILA_MOV, MAX_FILA_MOV + 1):
        for c in range(2, 10):
            ws.cell(r, c).value = None
    for i, f in enumerate(filas):
        r = PRIMERA_FILA_MOV + i
        ws.cell(r, 2, f["fecha"]).number_format = "dd/mm/yyyy"
        ws.cell(r, 3, f'=IF(B{r}="","",TEXT(B{r},"mmm-yy"))')
        for col, clave in ((4, "fuente"), (5, "descripcion"), (6, "moneda")):
            ws.cell(r, col, f[clave]).font = Font(name=FN, color=INK, size=10)
        ws.cell(r, 7, f["monto"]).font = Font(name=FN, color=BLUE)
        ws.cell(r, 8, f'=IF(OR(G{r}="",F{r}=""),"",IF(F{r}="ARS",G{r},G{r}*USDT_ARS))')
        ws.cell(r, 9, f'=IF(OR(G{r}="",F{r}=""),"",IF(F{r}="USDT",G{r},G{r}/USDT_ARS))')


def agregar_gasto(datos):
    fila = {
        "fecha": _fecha(datos.get("fecha")),
        "categoria": _texto(datos.get("categoria")) or "Otros",
        "descripcion": _texto(datos.get("descripcion")),
        "medio": _texto(datos.get("medio")),
        "tipo": _texto(datos.get("tipo")) or "Variable",
        "moneda": _texto(datos.get("moneda")) or "ARS",
        "monto": abs(_numero(datos.get("monto"))),
    }
    with editar() as wb:
        ws = wb["Gastos"]
        cols = {"fecha": 2, "categoria": 4, "descripcion": 5, "medio": 6,
                "tipo": 7, "moneda": 8, "monto": 9}
        filas = _leer_filas(ws, cols)
        filas.append(fila)
        _escribir_gastos(ws, filas)
    return f"Gasto de {fila['monto']:,.0f} {fila['moneda']} guardado."


def agregar_gastos(lista):
    """Alta de varios gastos de una sola vez (importación de MercadoPago).

    Hacerlo en una sola operación evita un backup y una reescritura completa
    de la hoja por cada fila del CSV.
    """
    nuevas = []
    for d in lista:
        nuevas.append({
            "fecha": _fecha(d.get("fecha")),
            "categoria": _texto(d.get("categoria")) or "Otros",
            "descripcion": _texto(d.get("descripcion")),
            "medio": _texto(d.get("medio")) or "MercadoPago",
            "tipo": _texto(d.get("tipo")) or "Variable",
            "moneda": _texto(d.get("moneda")) or "ARS",
            "monto": abs(_numero(d.get("monto"))),
        })
    if not nuevas:
        return "No había nada para cargar."
    with editar() as wb:
        ws = wb["Gastos"]
        cols = {"fecha": 2, "categoria": 4, "descripcion": 5, "medio": 6,
                "tipo": 7, "moneda": 8, "monto": 9}
        filas = _leer_filas(ws, cols)
        filas.extend(nuevas)
        _escribir_gastos(ws, filas)
    return f"{len(nuevas)} gastos cargados."


def agregar_ingreso(datos):
    fila = {
        "fecha": _fecha(datos.get("fecha")),
        "fuente": _texto(datos.get("fuente")) or "Otros",
        "descripcion": _texto(datos.get("descripcion")),
        "moneda": _texto(datos.get("moneda")) or "ARS",
        "monto": abs(_numero(datos.get("monto"))),
    }
    with editar() as wb:
        ws = wb["Ingresos"]
        cols = {"fecha": 2, "fuente": 4, "descripcion": 5, "moneda": 6, "monto": 7}
        filas = _leer_filas(ws, cols)
        filas.append(fila)
        _escribir_ingresos(ws, filas)
    return f"Ingreso de {fila['monto']:,.0f} {fila['moneda']} guardado."


def borrar_movimiento(hoja, fila):
    """Borra una fila de Gastos o Ingresos y recompacta la hoja."""
    if hoja not in ("Gastos", "Ingresos"):
        raise ErrorCRM("Hoja inválida.")
    with editar() as wb:
        ws = wb[hoja]
        if hoja == "Gastos":
            cols = {"fecha": 2, "categoria": 4, "descripcion": 5, "medio": 6,
                    "tipo": 7, "moneda": 8, "monto": 9}
        else:
            cols = {"fecha": 2, "fuente": 4, "descripcion": 5, "moneda": 6, "monto": 7}
        filas = []
        for r in range(PRIMERA_FILA_MOV, MAX_FILA_MOV + 1):
            if ws.cell(r, 2).value is None:
                continue
            if r == int(fila):
                continue
            filas.append({n: ws.cell(r, c).value for n, c in cols.items()})
        if hoja == "Gastos":
            _escribir_gastos(ws, filas)
        else:
            _escribir_ingresos(ws, filas)
    return "Movimiento borrado."


# ---------------------------------------------------------------------------
# Escritura: aportes de inversión
# ---------------------------------------------------------------------------
def agregar_aporte(datos):
    fecha = _fecha(datos.get("fecha"))
    monto = abs(_numero(datos.get("monto")))
    moneda = _texto(datos.get("moneda")) or "ARS"
    nota = _texto(datos.get("nota"))

    with editar() as wb:
        ws = wb["Inversiones"]
        sec = localizar_secciones(ws)
        if "aportes" not in sec:
            raise ErrorCRM("No encontré la tabla de aportes en la hoja Inversiones.")
        ini, fin = sec["aportes"]
        fila = _primera_libre(ws, ini, fin)
        if fila is None:
            raise ErrorCRM("La tabla de aportes se quedó sin filas libres.")
        ws.cell(fila, 2, fecha).number_format = "dd/mm/yyyy"
        ws.cell(fila, 4, monto)
        ws.cell(fila, 5, moneda)
        ws.cell(fila, 6, nota)
    return f"Aporte de {monto:,.0f} {moneda} guardado."


# ---------------------------------------------------------------------------
# Escritura: posiciones de Bull Market
#
# Se reescribe el bloque entero con la lista que manda el CRM, así se pueden
# editar, agregar y borrar posiciones en una sola operación.
# ---------------------------------------------------------------------------
def guardar_posiciones(posiciones):
    limpias = []
    for p in posiciones:
        activo = _texto(p.get("activo"))
        if not activo:
            continue
        limpias.append({
            "activo": activo,
            "tipo": _texto(p.get("tipo")) or "Acción AR",
            "fechaCompra": _fecha(p.get("fechaCompra")) if p.get("fechaCompra") else None,
            "cantidad": _numero(p.get("cantidad"), "cantidad"),
            "precioCompra": _numero(p.get("precioCompra"), "precio de compra"),
            "precioActual": _numero(p.get("precioActual"), "precio actual"),
            "meta": _texto(p.get("meta")),
        })

    with editar() as wb:
        ws = wb["Inversiones"]
        sec = localizar_secciones(ws)
        if "bull_market" not in sec:
            raise ErrorCRM("No encontré el bloque de Bull Market en la hoja Inversiones.")
        ini, fin = sec["bull_market"]
        disponibles = fin - ini + 1
        if len(limpias) > disponibles:
            raise ErrorCRM(
                f"Entran {disponibles} posiciones y mandaste {len(limpias)}. "
                "Hay que agrandar el bloque de Bull Market."
            )
        for r in range(ini, fin + 1):
            for c in (2, 3, 4, 5, 6, 7, 12):
                ws.cell(r, c).value = None
        for i, p in enumerate(limpias):
            r = ini + i
            ws.cell(r, 2, p["activo"])
            ws.cell(r, 3, p["tipo"])
            if p["fechaCompra"]:
                ws.cell(r, 4, p["fechaCompra"]).number_format = "dd/mm/yyyy"
            ws.cell(r, 5, p["cantidad"])
            ws.cell(r, 6, p["precioCompra"])
            ws.cell(r, 7, p["precioActual"])
            if p["meta"]:
                ws.cell(r, 12, p["meta"])
    return f"{len(limpias)} posiciones de Bull Market guardadas."


# ---------------------------------------------------------------------------
# Escritura: metas y ventas de mercadería
# ---------------------------------------------------------------------------
def agregar_meta(datos):
    nombre = _texto(datos.get("meta"))
    if not nombre:
        raise ErrorCRM("Falta el nombre de la meta.")
    objetivo = abs(_numero(datos.get("objetivo"), "objetivo"))

    with editar() as wb:
        ws = wb["Metas"]
        fila = _primera_libre(ws, 5, ws.max_row)
        if fila is None:
            raise ErrorCRM("La hoja Metas se quedó sin filas libres.")
        ws.cell(fila, 2, nombre)
        ws.cell(fila, 3, _texto(datos.get("categoria")) or "Otro")
        if datos.get("fechaObjetivo"):
            ws.cell(fila, 4, _fecha(datos["fechaObjetivo"])).number_format = "dd/mm/yyyy"
        ws.cell(fila, 5, _texto(datos.get("moneda")) or "ARS")
        ws.cell(fila, 6, objetivo)
        if datos.get("ahorradoManual"):
            ws.cell(fila, 7, _numero(datos["ahorradoManual"], "ahorrado"))
    return f'Meta "{nombre}" guardada.'


def agregar_venta(datos):
    fecha = _fecha(datos.get("fecha"))
    cantidad = abs(_numero(datos.get("cantidad"), "cantidad"))
    precio = abs(_numero(datos.get("precioVenta"), "precio de venta"))
    producto = _texto(datos.get("producto")) or "Otro"

    with editar() as wb:
        ws = wb["Inversiones"]
        sec = localizar_secciones(ws)
        if "ventas" not in sec:
            raise ErrorCRM("No encontré el registro de ventas en la hoja Inversiones.")
        ini, fin = sec["ventas"]
        fila = _primera_libre(ws, ini, fin)
        if fila is None:
            raise ErrorCRM("El registro de ventas se quedó sin filas libres.")
        ws.cell(fila, 2, fecha).number_format = "dd/mm/yyyy"
        ws.cell(fila, 3, producto)
        ws.cell(fila, 4, cantidad)
        ws.cell(fila, 5, precio)
    return f"Venta de {cantidad:g} unidades guardada."


# ---------------------------------------------------------------------------
# Escritura: cotización (Config)
# ---------------------------------------------------------------------------
def guardar_cotizacion(usdt_ars=None, dolar_blue=None):
    with editar() as wb:
        ws = wb["Config"]
        if usdt_ars:
            ws["C5"] = _numero(usdt_ars, "dólar USDT")
        if dolar_blue:
            ws["C6"] = _numero(dolar_blue, "dólar blue")
    return "Cotización actualizada."
