"""
Importar la tenencia de Bull Market al CRM, a partir del reporte de
"Cuenta Corriente" (CSV o Excel) — es lo único que Bull Market exporta para
este tipo de cuenta; no hay un reporte de "tenencia" aparte.

La Cuenta Corriente es un LIBRO DE MOVIMIENTOS (una fila por cada compra,
venta, dividendo, transferencia, caución...), no una foto de lo que se tiene
hoy. Por eso este módulo reconstruye la tenencia sumando: cada fila con una
Especie real (no "VARIAS", que es como Bull Market marca las cauciones) y
Cantidad distinta de cero suma o resta a esa posición — positivo compra,
negativo venta. Esa misma regla, sin casos especiales por tipo de operación,
ya cubre compras, ventas, rescates de fondos y amortizaciones de bonos,
porque todas mueven la Cantidad de una Especie real; lo que no mueve
Cantidad (dividendos, transferencias, retenciones) queda afuera solo.

El precio de compra se saca del promedio ponderado de las filas que suman
(las compras). El precio actual no existe en un libro de movimientos —se
usa el precio de la operación más reciente de cada especie como mejor
estimación, aclarado como tal en el resultado para que se sepa que no es un
precio en vivo.
"""
import csv
import io
import re
import unicodedata
from datetime import date, datetime

from openpyxl import load_workbook


class ErrorBM(Exception):
    """Error previsible al importar, para mostrar tal cual en pantalla."""


PALABRAS_CLAVE = {
    "fecha": ["operado", "fecha operado", "fecha"],
    "ticker": ["especie", "ticker", "simbolo"],
    "cantidad": ["cantidad"],
    "precio": ["precio"],
    "saldo": ["saldo"],
}
# "saldo" no es obligatorio para reconocer el archivo (se sigue pudiendo
# reconstruir la tenencia sin él) — por eso no está en ORDEN_DETECCION, que
# es lo que se exige como mínimo.
ORDEN_DETECCION = ["fecha", "ticker", "cantidad", "precio", "saldo"]
COLUMNAS_MINIMAS = ["fecha", "ticker", "cantidad", "precio"]
FILAS_A_ESCANEAR_HEADER = 15

# Tickers de CEDEARs (empresas extranjeras que cotizan en pesos en Bull
# Market) conocidos, para sugerir el "Tipo" de la posición nueva. Es sólo un
# valor por defecto — se edita a mano en la tabla si hace falta.
CEDEARS_CONOCIDOS = {
    "MSFT", "NVDA", "AAPL", "GOOGL", "GOOG", "AMZN", "TSLA", "META", "NFLX",
    "DIS", "KO", "PBR", "BABA", "PYPL", "V", "MA", "JNJ", "PFE", "XOM",
    "WMT", "NKE", "SBUX", "UBER", "ABNB", "COIN", "MELI",
}
# Letras/bonos del Tesoro argentino: "S" + 2 dígitos + letra de mes + dígito
# de año (ej. S27F6, S10N5).
RE_BONO_AR = re.compile(r"^S\d{2}[A-Z]\d$")


def _normalizar(s):
    s = str(s or "").strip().lower()
    s = unicodedata.normalize("NFKD", s)
    return "".join(c for c in s if not unicodedata.combining(c))


def _tipo_sugerido(ticker):
    t = ticker.strip().upper()
    if RE_BONO_AR.match(t):
        return "Bono AR"
    if t in CEDEARS_CONOCIDOS:
        return "CEDEAR"
    return "Acción AR"


def _parsear_fecha(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v or "").strip()
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parsear_numero(v):
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v or "").strip()
    if not s:
        return None
    i_coma, i_punto = s.rfind(","), s.rfind(".")
    if i_coma > -1 and i_punto > -1:
        s = s.replace(".", "").replace(",", ".") if i_coma > i_punto else s.replace(",", "")
    elif i_coma > -1:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _leer_csv(data):
    texto = None
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            texto = data.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if texto is None:
        raise ErrorBM("No pude leer el archivo — probá exportarlo de nuevo.")
    muestra = texto[:4096]
    try:
        sep = csv.Sniffer().sniff(muestra, delimiters=",;\t").delimiter
    except csv.Error:
        sep = ";" if muestra.count(";") > muestra.count(",") else ","
    filas = list(csv.reader(io.StringIO(texto), delimiter=sep))
    return [fila for fila in filas if any((c or "").strip() for c in fila)]


def _leer_excel(data):
    # Ojo: SIN read_only=True. Algunos exports de Bull Market traen mal la
    # etiqueta interna de rango de datos del .xlsx (dice que el archivo
    # termina en la columna E cuando en realidad sigue hasta la J) — en modo
    # read_only openpyxl confía en esa etiqueta y corta columnas de verdad.
    # Cargando normal, escanea las celdas reales y no pasa.
    wb = load_workbook(io.BytesIO(data), data_only=True)
    mejor, filas_mejor = None, -1
    for nombre in wb.sheetnames:
        ws = wb[nombre]
        filas = [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
                 for r in range(1, ws.max_row + 1)]
        filas = [f for f in filas if any(c is not None and str(c).strip() for c in f)]
        if len(filas) > filas_mejor:
            mejor, filas_mejor = filas, len(filas)
    wb.close()
    return mejor or []


def _detectar_header(filas):
    mejor_idx, mejor_deteccion, mejor_puntaje = None, {}, 0
    tope = min(FILAS_A_ESCANEAR_HEADER, len(filas))
    for i in range(tope):
        fila_norm = [_normalizar(c) for c in filas[i]]
        deteccion, usadas = {}, set()
        for categoria in ORDEN_DETECCION:
            for j, celda in enumerate(fila_norm):
                if j in usadas or not celda:
                    continue
                if any(clave in celda for clave in PALABRAS_CLAVE[categoria]):
                    deteccion[categoria] = j
                    usadas.add(j)
                    break
        puntaje = len(deteccion)
        if puntaje > mejor_puntaje:
            mejor_idx, mejor_deteccion, mejor_puntaje = i, deteccion, puntaje
    return mejor_idx, mejor_deteccion, mejor_puntaje


def analizar(nombre_archivo, data):
    """Lee la Cuenta Corriente y devuelve la tenencia ya reconstruida.

    No escribe nada al Excel: crm_ui.html fusiona lo devuelto con las
    posiciones que ya había y arma la tabla editable de siempre; recién al
    apretar "Guardar posiciones" se toca el archivo.
    """
    ext = (nombre_archivo.rsplit(".", 1)[-1] if "." in nombre_archivo else "").lower()
    if ext in ("xlsx", "xlsm", "xls"):
        filas = _leer_excel(data)
    elif ext == "csv":
        filas = _leer_csv(data)
    else:
        raise ErrorBM("Subí un CSV o un Excel (.csv, .xlsx) — ese formato no lo reconozco.")

    if not filas:
        raise ErrorBM("El archivo está vacío.")

    idx_header, det, puntaje = _detectar_header(filas)
    faltan = [c for c in COLUMNAS_MINIMAS if c not in det]
    if idx_header is None or faltan:
        return {
            "reconocido": False,
            "encabezados": filas[idx_header if idx_header is not None else 0],
            "posiciones": [],
            "advertencias": [],
        }

    # --------------------------------------------------------------------
    # Reconstruir la tenencia: para cada Especie real (no "VARIAS", que es
    # como Bull Market anota las cauciones/repos), se acumula la cantidad
    # movida en cada fila y se pondera el precio de las filas que suman
    # (las compras) para sacar el precio de compra promedio.
    # --------------------------------------------------------------------
    acumulado = {}  # ticker -> {cantidad, costo_total, cant_compra, ultima_fecha, ultimo_precio, movimientos}
    for fila in filas[idx_header + 1:]:
        if det["ticker"] >= len(fila) or det["cantidad"] >= len(fila):
            continue
        ticker = str(fila[det["ticker"]] or "").strip().upper()
        if not ticker or ticker == "VARIAS":
            continue
        cantidad = _parsear_numero(fila[det["cantidad"]])
        if not cantidad:
            continue
        precio = _parsear_numero(fila[det["precio"]]) if det["precio"] < len(fila) else None
        fecha = _parsear_fecha(fila[det["fecha"]]) if det["fecha"] < len(fila) else None

        reg = acumulado.setdefault(ticker, {
            "cantidad": 0.0, "costo_total": 0.0, "cant_compra": 0.0,
            "primera_compra": None, "ultima_fecha": None, "ultimo_precio": None,
            "movimientos": 0,
        })
        reg["cantidad"] += cantidad
        reg["movimientos"] += 1
        if cantidad > 0 and precio:
            reg["costo_total"] += cantidad * precio
            reg["cant_compra"] += cantidad
            if reg["primera_compra"] is None or (fecha and fecha < reg["primera_compra"]):
                reg["primera_compra"] = fecha
        if fecha and (reg["ultima_fecha"] is None or fecha >= reg["ultima_fecha"]):
            reg["ultima_fecha"] = fecha
            reg["ultimo_precio"] = precio

    posiciones, advertencias = [], []
    EPS = 1e-6
    for ticker, reg in sorted(acumulado.items()):
        if reg["cantidad"] <= EPS:
            if reg["cantidad"] < -EPS:
                advertencias.append(
                    f'{ticker}: la cantidad neta de este archivo dio negativa — probablemente '
                    "tenías compras de antes del rango de fechas de este reporte. No se importó; "
                    "cargala a mano si hace falta."
                )
            continue
        precio_compra = (reg["costo_total"] / reg["cant_compra"]) if reg["cant_compra"] > EPS else reg["ultimo_precio"]
        posiciones.append({
            "activo": ticker,
            "tipo": _tipo_sugerido(ticker),
            "fechaCompra": (reg["primera_compra"] or reg["ultima_fecha"] or date.today()).isoformat(),
            "cantidad": round(reg["cantidad"], 6),
            "precioCompra": round(precio_compra, 6) if precio_compra else None,
            "precioActual": round(reg["ultimo_precio"], 6) if reg["ultimo_precio"] else None,
            "movimientos": reg["movimientos"],
        })

    if not posiciones and not advertencias:
        advertencias.append("No encontré ninguna operación de compra/venta de un activo real en este archivo.")

    # --------------------------------------------------------------------
    # Efectivo disponible: a diferencia de las posiciones de arriba (que se
    # reconstruyen sumando SOLO lo que aparece en este archivo, y por eso
    # pueden quedar cortas si hay historia previa), el Saldo es el balance
    # de caja acumulado que Bull Market ya viene arrastrando fila a fila —
    # el de la ÚLTIMA fila del archivo es, ese sí, el saldo real y
    # actualizado a la fecha del reporte. Por eso el efectivo se puede
    # actualizar siempre con este valor, a diferencia del resto.
    # --------------------------------------------------------------------
    efectivo = None
    if "saldo" in det:
        for fila in reversed(filas[idx_header + 1:]):
            if det["saldo"] >= len(fila):
                continue
            saldo = _parsear_numero(fila[det["saldo"]])
            if saldo is not None:
                fecha_saldo = _parsear_fecha(fila[det["fecha"]]) if det["fecha"] < len(fila) else None
                efectivo = {
                    "cantidad": round(saldo, 2),
                    "fecha": (fecha_saldo or date.today()).isoformat(),
                }
                break

    return {
        "reconocido": True,
        "posiciones": posiciones,
        "advertencias": advertencias,
        "encabezados": filas[idx_header],
        "efectivo": efectivo,
    }
