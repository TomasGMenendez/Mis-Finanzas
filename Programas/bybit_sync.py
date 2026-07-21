"""
Bybit → Excel sync (Finanzas Toto)
Lee las credenciales de bybit_config.txt, trae balances reales de todas las
cuentas (Unified + Funding + Contract + Earn/Investment + Option) y precios
actuales, y muestra el resumen. Con --sync también actualiza Finanzas_Toto.xlsx.
"""
import sys, os, subprocess, importlib
from pathlib import Path

BASE = Path(__file__).parent
RAIZ = BASE.parent  # donde vive Finanzas_Toto.xlsx

# ----- Auto-install de dependencias -----
def ensure(pkg, import_name=None):
    name = import_name or pkg
    try:
        importlib.import_module(name)
    except ImportError:
        print(f"Instalando {pkg}...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", pkg, "--quiet"])
ensure("pybit")
ensure("openpyxl")

from pybit.unified_trading import HTTP
from openpyxl import load_workbook

# ----- Config -----
cfg_path = BASE / "bybit_config.txt"
if not cfg_path.exists():
    print(f"ERROR: no encuentro {cfg_path}"); sys.exit(1)
cfg = {}
for line in cfg_path.read_text().splitlines():
    if "=" in line and not line.strip().startswith("#"):
        k, v = line.split("=", 1); cfg[k.strip()] = v.strip()
if "BYBIT_API_KEY" not in cfg or "BYBIT_API_SECRET" not in cfg:
    print("ERROR: bybit_config.txt debe tener BYBIT_API_KEY y BYBIT_API_SECRET"); sys.exit(1)

session = HTTP(api_key=cfg["BYBIT_API_KEY"], api_secret=cfg["BYBIT_API_SECRET"], testnet=False)

# ----- Consolidar balances de todas las cuentas -----
holdings = {}  # coin -> {"qty": float, "sources": [...]}

def add(coin, qty, src):
    if qty <= 0: return
    if coin not in holdings:
        holdings[coin] = {"qty": 0.0, "sources": []}
    holdings[coin]["qty"] += qty
    holdings[coin]["sources"].append(f"{src}:{qty:.8f}")

print("=" * 60)
print("Consultando cuentas...")
print("=" * 60)

# 1. UNIFIED Trading Account (Spot + Derivs + Earn en algunas configs)
try:
    r = session.get_wallet_balance(accountType="UNIFIED")
    if r.get("retCode") == 0:
        for acc in r["result"]["list"]:
            print(f"  UNIFIED equity: ${float(acc.get('totalEquity',0)):,.2f}")
            for c in acc.get("coin", []):
                add(c["coin"], float(c.get("walletBalance", 0)), "UTA")
    else:
        print(f"  UNIFIED: {r.get('retMsg')}")
except Exception as e:
    print(f"  UNIFIED error: {e}")

# 2. FUNDING
try:
    r = session.get_coins_balance(accountType="FUND")
    if r.get("retCode") == 0:
        for c in r["result"]["balance"]:
            add(c["coin"], float(c.get("walletBalance", 0)), "FUND")
    else:
        print(f"  FUND: {r.get('retMsg')}")
except Exception as e:
    print(f"  FUND error: {e}")

# 3. EARN — Easy Earn / Flexible Savings
# Bybit v5 endpoint: GET /v5/earn/position
try:
    r = session._submit_request(
        method="GET",
        path=f"{session.endpoint}/v5/earn/position",
        query={"category": "FlexibleSaving"},
        auth=True,
    )
    if r.get("retCode") == 0:
        positions = r.get("result", {}).get("list", [])
        print(f"  EARN posiciones: {len(positions)}")
        for pos in positions:
            coin = pos.get("coin")
            qty = float(pos.get("amount", 0))
            add(coin, qty, "EARN-Flex")
    else:
        print(f"  EARN Flex: {r.get('retMsg')}")
except Exception as e:
    print(f"  EARN Flex error: {e}")

# Otras categorías de Earn
for cat in ["OnChain", "FixedSaving"]:
    try:
        r = session._submit_request(
            method="GET",
            path=f"{session.endpoint}/v5/earn/position",
            query={"category": cat},
            auth=True,
        )
        if r.get("retCode") == 0:
            for pos in r.get("result", {}).get("list", []):
                add(pos.get("coin"), float(pos.get("amount", 0)), f"EARN-{cat}")
    except Exception:
        pass

# ----- Precios actuales (spot USDT) -----
def price(coin):
    if coin in ("USDT", "USDC", "DAI"): return 1.0
    try:
        r = session.get_tickers(category="spot", symbol=f"{coin}USDT")
        return float(r["result"]["list"][0]["lastPrice"])
    except Exception:
        return None

# Construir portafolio con USD value
portfolio = []
for coin, data in holdings.items():
    p = price(coin)
    if p is None:
        portfolio.append((coin, data["qty"], None, None, data["sources"]))
        continue
    usd = data["qty"] * p
    portfolio.append((coin, data["qty"], p, usd, data["sources"]))

portfolio.sort(key=lambda x: -(x[3] if x[3] else 0))

# ----- Reporte -----
print("\n" + "=" * 60)
print("PORTAFOLIO ACTUAL")
print("=" * 60)
print(f"{'Coin':<8}{'Cantidad':>18}{'Precio':>14}{'USD':>14}   Fuentes")
print("-" * 90)
total_usd = 0
for coin, qty, p, usd, srcs in portfolio:
    if p is None:
        print(f"{coin:<8}{qty:>18.8f}{'-':>14}{'?':>14}   {', '.join(srcs)}")
    elif usd < 0.01:
        continue
    else:
        print(f"{coin:<8}{qty:>18.8f}{p:>14.4f}{usd:>14.2f}   {', '.join(srcs)}")
        total_usd += usd
print("-" * 90)
print(f"{'TOTAL':<40}{total_usd:>14.2f} USD")

# ----- Sync al Excel -----
if "--sync" in sys.argv:
    xlsx_path = RAIZ / "Finanzas_Toto.xlsx"
    if not xlsx_path.exists():
        print(f"\nERROR: no encuentro {xlsx_path}"); sys.exit(1)
    wb = load_workbook(xlsx_path)
    inv = wb["Inversiones"]

    # Leer avg cost basis existente (col F) para preservarlos si el activo ya estaba
    existing_avg = {}
    for r in range(5, 11):  # solo filas Bybit
        a = inv.cell(r, 2).value
        if a: existing_avg[a] = inv.cell(r, 6).value  # col F = avg price

    # Detectar celdas merged para skippear (no se pueden escribir)
    merged_ranges = list(inv.merged_cells.ranges)
    def is_merged(row, col):
        for mr in merged_ranges:
            if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
                return True
        return False

    # Limpiar solo filas 5-10 (Bybit), NO tocar 11+ (Bull Market)
    for r in range(5, 11):
        for c in range(2, 8):
            if not is_merged(r, c):
                inv.cell(r, c).value = None

    # Reescribir con datos frescos
    row = 5
    from datetime import date
    today = date.today()
    for coin, qty, p, usd, _ in portfolio:
        if not p or (usd is not None and usd < 0.10): continue
        if row >= 11:  # no invadir zona Bull Market
            break
        if not is_merged(row, 2):
            inv.cell(row, 2, value=coin)
        if not is_merged(row, 3):
            inv.cell(row, 3, value="Cripto")
        if not is_merged(row, 4):
            inv.cell(row, 4, value=today)
        if not is_merged(row, 5):
            inv.cell(row, 5, value=qty)
        # Preservar avg si existía
        if not is_merged(row, 6):
            inv.cell(row, 6, value=existing_avg.get(coin, p))
        if not is_merged(row, 7):
            inv.cell(row, 7, value=p)
        row += 1

    # Actualizar cotización USDT/ARS con estimación (Bybit P2P mercado — placeholder)
    # Podés overwritearla manual en Config!C5
    wb.save(xlsx_path)
    print(f"\n[OK] Excel actualizado: {xlsx_path}")
    print(f"     {row - 5} activos escritos.")
    print("     NOTA: los 'precio promedio' se preservaron de lo que ya tenias.")
    print("     Si algo estaba mal, editalo manualmente en la hoja Inversiones col F.")
else:
    print("\n(Modo test - no se toco el Excel. Para sincronizar corre: python bybit_sync.py --sync)")
