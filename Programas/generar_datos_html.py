"""
Finanzas_Toto.xlsx -> datos para el Dashboard.

Lee Finanzas_Toto.xlsx (hojas Gastos, Ingresos, Inversiones, Metas, Config) y
arma el mismo dict de datos que consume "Mi Dashboard de Finanzas.html"
(la variable EMBEDDED_DATA).

Se usa de dos formas:
  - Como script (ver abajo, "python generar_datos_html.py"): reemplaza
    EMBEDDED_DATA dentro del Dashboard y su copia para publicar, así al
    abrirlo con doble clic (sin server, sin conexión en vivo) ya muestra los
    datos reales tal como estaban en el Excel al correr el script.
  - Importado desde crm_server.py (función construir_datos): el CRM sirve el
    Dashboard con datos en vivo, recalculados en cada carga de la página, sin
    tocar ningún archivo en disco.
"""
import sys, os, subprocess, json, re, shutil, tempfile
from pathlib import Path
from datetime import datetime, date

BASE = Path(__file__).parent
RAIZ = BASE.parent  # donde viven Finanzas_Toto.xlsx y el Dashboard

def ensure(pkg):
    try:
        __import__(pkg)
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", pkg, "--quiet"])
ensure("openpyxl")

from openpyxl import load_workbook


def cell(ws, r, c):
    return ws.cell(r, c).value


def num(v):
    if isinstance(v, (int, float)):
        return v
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0


def fecha_str(v):
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    return None


def construir_datos(xlsx_path):
    """Lee el Excel y devuelve el dict de datos que consume el Dashboard.

    No modifica nada — ni el Excel ni ningún archivo del Dashboard. Cada
    llamada lee el estado actual del archivo desde cero.
    """
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        raise FileNotFoundError(f"No encuentro {xlsx_path}")

    # El Excel puede estar abierto en la app de Excel o sincronizando con
    # OneDrive, lo que bloquea la lectura directa (incluso para copiar el
    # archivo). Se reintenta con una pausa breve porque casi siempre se
    # libera en segundos. Se copia a un directorio temporal propio (no un
    # nombre fijo) para que dos lecturas en simultáneo — ej. el CRM sirviendo
    # el Dashboard mientras alguien corre el script a mano — no choquen.
    with tempfile.TemporaryDirectory(prefix="finanzas_toto_") as tmpdir:
        tmp_path = Path(tmpdir) / "libro.xlsx"
        last_err = None
        for attempt in range(5):
            try:
                shutil.copyfile(xlsx_path, tmp_path)
                last_err = None
                break
            except PermissionError as e:
                last_err = e
                if attempt < 4:
                    import time
                    time.sleep(2)
        if last_err:
            raise PermissionError(
                f"No pude leer {xlsx_path} — está abierto en Excel o sincronizando con OneDrive."
            )
        wb = load_workbook(tmp_path, data_only=True)

        gastos = []
        ws = wb["Gastos"]
        for r in range(5, ws.max_row + 1):
            f = fecha_str(cell(ws, r, 2))
            monto = num(cell(ws, r, 9))
            if not f or not monto:
                continue
            gastos.append({
                "fecha": f,
                "cat": (cell(ws, r, 4) or "").strip() if isinstance(cell(ws, r, 4), str) else (cell(ws, r, 4) or ""),
                "desc": cell(ws, r, 5) or "",
                "medio": cell(ws, r, 6) or "",
                "tipo": cell(ws, r, 7) or "",
                "moneda": cell(ws, r, 8) or "ARS",
                "monto": monto,
            })

        ingresos = []
        ws = wb["Ingresos"]
        for r in range(5, ws.max_row + 1):
            f = fecha_str(cell(ws, r, 2))
            monto = num(cell(ws, r, 7))
            if not f or not monto:
                continue
            ingresos.append({
                "fecha": f,
                "fuente": cell(ws, r, 4) or "",
                "desc": cell(ws, r, 5) or "",
                "moneda": cell(ws, r, 6) or "ARS",
                "monto": monto,
            })

        inversiones = []
        ws = wb["Inversiones"]
        skip_re = re.compile(r"^──|^Subtotal|^RESUMEN|^Secci[oó]n|^Bybit|^Bull Market|^TOTAL", re.I)
        # Debajo del portafolio real (cripto + Bull Market) puede haber otras
        # secciones con otro formato de columnas — hoy "RESUMEN CONSOLIDADO" e
        # "INVERSIONES FÍSICAS" (ej. mercadería para reventa). No son holdings
        # de portafolio, así que hay que dejar de leer filas apenas aparece
        # cualquiera de esas secciones, en vez de solo saltear esa fila puntual.
        stop_re = re.compile(r"^RESUMEN CONSOLIDADO|^── INVERSIONES", re.I)
        # Columna "Meta" (vincula un activo a un objetivo de la hoja Metas): se
        # busca por texto de encabezado en vez de posición fija, por si se
        # agregan o reordenan columnas más adelante.
        meta_col = None
        for c in range(1, ws.max_column + 1):
            if str(cell(ws, 4, c) or "").strip().lower() == "meta":
                meta_col = c
                break
        for r in range(5, ws.max_row + 1):
            activo = cell(ws, r, 2)
            if activo and stop_re.match(str(activo)):
                break
            if not activo or skip_re.match(str(activo)):
                continue
            cantidad = num(cell(ws, r, 5))
            if not cantidad:
                continue
            inversiones.append({
                "activo": str(activo).strip(),
                "tipo": cell(ws, r, 3) or "",
                "cantidad": cantidad,
                "precioCompra": num(cell(ws, r, 6)),
                "precioActual": num(cell(ws, r, 7)),
                "meta": (str(cell(ws, r, meta_col)).strip() if meta_col and cell(ws, r, meta_col) else ""),
            })

        # Inversiones físicas (mercadería para reventa, ej. "Bolsas comida
        # perro"): viven en su propia sub-sección de la hoja Inversiones, con
        # otro formato de columnas. Se calcula todo desde los datos crudos
        # (cantidad, costo, precio de venta) en vez de confiar en las
        # fórmulas de Excel — esas quedan cacheadas desactualizadas si el
        # archivo no se reabre en Excel para recalcular (mismo problema que
        # tuvimos con la hoja Ingresos).
        ws = wb["Inversiones"]
        productos_fisicos = []
        ventas_fisicas = []
        inv_fis_start = None
        for r in range(1, ws.max_row + 1):
            v = cell(ws, r, 2)
            if v and re.match(r"^── INVERSIONES", str(v), re.I):
                inv_fis_start = r
                break
        if inv_fis_start:
            # Sub-tabla de compras: busca el header "Producto" / "Fecha
            # compra" y lee filas hasta la próxima fila vacía o de sección.
            r = inv_fis_start + 1
            while r <= ws.max_row and not (cell(ws, r, 2) == "Producto" and cell(ws, r, 3) == "Fecha compra"):
                r += 1
            r += 1
            while r <= ws.max_row:
                producto = cell(ws, r, 2)
                if not producto or str(producto).strip() == "":
                    break
                cant = num(cell(ws, r, 4))
                costo = num(cell(ws, r, 5))
                if cant:
                    productos_fisicos.append({
                        "producto": str(producto).strip(),
                        "fechaCompra": fecha_str(cell(ws, r, 3)),
                        "cantidad": cant,
                        "costoUnitario": costo,
                    })
                r += 1

            # Sub-tabla "Registro de ventas": busca el header "Fecha venta" y
            # lee filas hasta "TOTAL VENTAS" o una fila vacía.
            while r <= ws.max_row and cell(ws, r, 2) != "Fecha venta":
                if cell(ws, r, 2) and re.match(r"^TOTAL", str(cell(ws, r, 2)), re.I):
                    break
                r += 1
            r += 1
            while r <= ws.max_row:
                fv = cell(ws, r, 2)
                if fv and re.match(r"^TOTAL", str(fv), re.I):
                    break
                cantidad_v = num(cell(ws, r, 4))
                precio_v = num(cell(ws, r, 5))
                if isinstance(fv, (datetime, date)) and cantidad_v and precio_v:
                    ventas_fisicas.append({
                        "fecha": fecha_str(fv),
                        "cantidad": cantidad_v,
                        "precioVenta": precio_v,
                    })
                elif fv is None and not cantidad_v and not precio_v:
                    break
                r += 1

        # Aportes a inversiones (tabla simple, aparte de Gastos): cada vez
        # que se le mete plata a una inversión sin desglosar por activo ni
        # precio. Vive en su propia sub-sección de la hoja Inversiones,
        # después de "REGISTRO DE VENTAS".
        ws = wb["Inversiones"]
        aportes_inversion = []
        ap_start = None
        for r in range(1, ws.max_row + 1):
            v = cell(ws, r, 2)
            if v and re.match(r"^── APORTES", str(v), re.I):
                ap_start = r
                break
        if ap_start:
            r = ap_start + 1
            while r <= ws.max_row and not (cell(ws, r, 2) == "Fecha" and cell(ws, r, 4) == "Monto"):
                r += 1
            r += 1
            while r <= ws.max_row:
                f = fecha_str(cell(ws, r, 2))
                monto = num(cell(ws, r, 4))
                if f and monto:
                    moneda = cell(ws, r, 5)
                    aportes_inversion.append({
                        "fecha": f,
                        "monto": monto,
                        "moneda": (moneda or "ARS").strip() if isinstance(moneda, str) else (moneda or "ARS"),
                        "nota": cell(ws, r, 6) or "",
                    })
                r += 1

        # Metas: objetivos de ahorro (ej. "Departamento", USDT 80.000).
        # "Ahorrado" es solo la parte manual (plata fuera del portafolio); lo
        # que aportan los activos vinculados (columna Meta de Inversiones) se
        # suma en el Dashboard, no acá, porque el precio en vivo no existe en
        # el Excel.
        metas = []
        ws = wb["Metas"]
        for r in range(5, ws.max_row + 1):
            nombre = cell(ws, r, 2)
            if not nombre:
                continue
            metas.append({
                "meta": str(nombre).strip(),
                "categoria": cell(ws, r, 3) or "",
                "fechaObjetivo": fecha_str(cell(ws, r, 4)),
                "moneda": (cell(ws, r, 5) or "ARS").strip() if isinstance(cell(ws, r, 5), str) else (cell(ws, r, 5) or "ARS"),
                "objetivo": num(cell(ws, r, 6)),
                "ahorradoManual": num(cell(ws, r, 7)),
            })

        ws = wb["Config"]
        usdt_ars = num(cell(ws, 5, 3)) or 1000
        blue_rate = num(cell(ws, 6, 3)) or usdt_ars

        wb.close()

    # Historial diario del portafolio (para el gráfico de evolución): lo
    # arma snapshot_portafolio.py corriendo solo todas las noches, no esta
    # función. Acá solo se lee ese archivo si existe — así no hace falta
    # tocar el Excel ni pedir precios en vivo.
    historial_path = BASE / "historial_portafolio.json"
    historial_portafolio = []
    if historial_path.exists():
        try:
            historial_portafolio = json.loads(historial_path.read_text(encoding="utf-8"))
        except Exception:
            historial_portafolio = []

    return {
        "gastos": gastos,
        "ingresos": ingresos,
        "inversiones": inversiones,
        "inversionesFisicas": {"productos": productos_fisicos, "ventas": ventas_fisicas},
        "aportesInversion": aportes_inversion,
        "metas": metas,
        "historialPortafolio": historial_portafolio,
        "config": {"usdtArs": usdt_ars, "blueRate": blue_rate},
    }


def _main():
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else RAIZ / "Finanzas_Toto.xlsx"
    try:
        data = construir_datos(xlsx_path)
    except FileNotFoundError as e:
        print(e)
        sys.exit(1)
    except PermissionError as e:
        print(e)
        print("Guardalo y cerralo (o esperá a que OneDrive termine de sincronizar) y volvé a correr el script.")
        sys.exit(1)

    html_path = RAIZ / "Mi Dashboard de Finanzas.html"
    html = html_path.read_text(encoding="utf-8")
    new_line = "var EMBEDDED_DATA = " + json.dumps(data, ensure_ascii=False, separators=(",", ":")) + ";"
    new_html, n = re.subn(r"var EMBEDDED_DATA = .*?;", lambda m: new_line, html, count=1)
    if n == 0:
        print("No encontré 'var EMBEDDED_DATA' en el Dashboard — no se modificó nada.")
        sys.exit(1)

    html_path.write_text(new_html, encoding="utf-8")
    print(f"Dashboard actualizado con datos reales: {len(data['gastos'])} gastos, "
          f"{len(data['ingresos'])} ingresos, {len(data['inversiones'])} inversiones, "
          f"{len(data['aportesInversion'])} aportes.")

    # Copia para publicar (docs/index.html, versión iPhone/PWA). Si todavía
    # no existe esa carpeta no pasa nada — este paso es opcional y no afecta
    # al Dashboard principal.
    docs_path = RAIZ / "docs" / "index.html"
    if docs_path.exists():
        docs_html = docs_path.read_text(encoding="utf-8")
        new_docs_html, n2 = re.subn(r"var EMBEDDED_DATA = .*?;", lambda m: new_line, docs_html, count=1)
        if n2:
            docs_path.write_text(new_docs_html, encoding="utf-8")
            print("Copia para el iPhone (docs/index.html) también actualizada.")
    print('Abrí "Mi Dashboard de Finanzas.html" haciendo doble clic para verlo con estos datos.')


if __name__ == "__main__":
    _main()
