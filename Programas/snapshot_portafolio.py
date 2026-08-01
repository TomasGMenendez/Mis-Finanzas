"""
Foto diaria del valor del portafolio (para el gráfico "Evolución del
portafolio" del Dashboard). Pensado para correr solo, todas las noches, vía
el Programador de tareas de Windows — no depende de que Toto abra el
Dashboard ni el Excel.

Qué hace:
  1. Lee la hoja Inversiones del Excel (si está abierto, no rompe: reintenta
     y si no puede, se lo salta por hoy — no es motivo para fallar todo).
  2. Busca precios en vivo (cripto en CoinGecko, acciones/CEDEARs/bonos
     argentinos en data912.com, dólar blue en dolarapi.com).
  3. Calcula el valor total del portafolio en pesos y lo guarda/actualiza en
     historial_portafolio.json (un renglón por día — no toca el Excel).
  4. Corre generar_datos_html.py para que el Dashboard (compu e iPhone)
     incluya ese historial.
  5. Sube los cambios a GitHub solo (sin pedir confirmación) para que el
     iPhone también lo vea.

Cualquier error queda anotado en snapshot_log.txt en vez de cortar en seco
(esto corre sin que nadie esté mirando la pantalla).
"""
import sys
import subprocess
import json
import re
import shutil
import tempfile
import time
from datetime import date, datetime
from pathlib import Path

BASE = Path(__file__).parent
RAIZ = BASE.parent
XLSX_PATH = RAIZ / "Finanzas_Toto.xlsx"
HIST_PATH = BASE / "historial_portafolio.json"
LOG_PATH = BASE / "snapshot_log.txt"


def log(msg):
    line = f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {msg}"
    print(line)
    with open(LOG_PATH, "a", encoding="utf-8") as f:
        f.write(line + "\n")


def ensure(pkg):
    try:
        __import__(pkg)
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", pkg, "--quiet"])


ensure("openpyxl")
ensure("requests")
from openpyxl import load_workbook
import requests

COINGECKO_IDS = {
    "BTC": "bitcoin", "ETH": "ethereum", "USDT": "tether", "XRP": "ripple",
    "ETC": "ethereum-classic", "BRETT": "based-brett", "SOL": "solana",
    "ADA": "cardano", "DOGE": "dogecoin", "DOT": "polkadot", "LTC": "litecoin",
    "AVAX": "avalanche-2", "LINK": "chainlink", "TRX": "tron", "SHIB": "shiba-inu",
    "BNB": "binancecoin", "USDC": "usd-coin", "MATIC": "matic-network", "PEPE": "pepe",
}


def norm_ticker(activo):
    return re.sub(r"[^A-Za-z0-9]", "", str(activo)).upper()


def read_inversiones_y_blue():
    tmp_path = Path(tempfile.gettempdir()) / "_snap_Finanzas_Toto.xlsx"
    ok = False
    for _ in range(5):
        try:
            shutil.copyfile(XLSX_PATH, tmp_path)
            ok = True
            break
        except PermissionError:
            time.sleep(2)
    if not ok:
        raise RuntimeError("Excel abierto/bloqueado, no se pudo leer.")

    wb = load_workbook(tmp_path, data_only=True)

    def cell(ws, r, c):
        return ws.cell(r, c).value

    def num(v):
        if isinstance(v, (int, float)):
            return v
        try:
            return float(v)
        except (TypeError, ValueError):
            return 0

    ws = wb["Inversiones"]
    skip_re = re.compile(r"^──|^Subtotal|^RESUMEN|^Secci[oó]n|^Bybit|^Bull Market|^TOTAL", re.I)
    stop_re = re.compile(r"^RESUMEN CONSOLIDADO|^── INVERSIONES", re.I)
    inversiones = []
    for r in range(5, ws.max_row + 1):
        activo = cell(ws, r, 2)
        if activo and stop_re.match(str(activo)):
            break
        if not activo or skip_re.match(str(activo)):
            continue
        cantidad = num(cell(ws, r, 5))
        if not cantidad:
            continue
        inversiones.append({
            "activo": str(activo).strip(),
            "tipo": cell(ws, r, 3) or "",
            "cantidad": cantidad,
            "precioActual": num(cell(ws, r, 7)),
        })

    ws = wb["Config"]
    blue_rate = num(cell(ws, 6, 3)) or num(cell(ws, 5, 3)) or 1000

    wb.close()
    tmp_path.unlink(missing_ok=True)
    return inversiones, blue_rate


def fetch_crypto_prices(tickers):
    ids = sorted(set(COINGECKO_IDS[t] for t in tickers if t in COINGECKO_IDS))
    if not ids:
        return {}
    try:
        r = requests.get(
            "https://api.coingecko.com/api/v3/coins/markets",
            params={"vs_currency": "usd", "ids": ",".join(ids)}, timeout=10,
        )
        r.raise_for_status()
        return {(c.get("symbol") or "").upper(): c.get("current_price") for c in r.json()}
    except Exception as e:
        log(f"Aviso: no se pudieron traer precios cripto ({e}). Sigo con los del Excel.")
        return {}


def fetch_live_ar_prices():
    out = {}
    for path in ("arg_stocks", "arg_cedears", "arg_bonds"):
        try:
            r = requests.get(f"https://data912.com/live/{path}", timeout=10)
            r.raise_for_status()
            for it in r.json():
                sym = (it.get("symbol") or "").upper()
                c = it.get("c")
                if sym and isinstance(c, (int, float)):
                    out[sym] = c
        except Exception as e:
            log(f"Aviso: no se pudieron traer precios de {path} ({e}). Sigo con los del Excel.")
    return out


def fetch_blue_rate(fallback):
    try:
        r = requests.get("https://dolarapi.com/v1/dolares/blue", timeout=10)
        r.raise_for_status()
        d = r.json()
        return d.get("venta") or d.get("compra") or fallback
    except Exception as e:
        log(f"Aviso: no se pudo traer el dólar blue en vivo ({e}). Uso el del Excel.")
        return fallback


def compute_portfolio_value_ars(inversiones, blue_rate):
    crypto_tks = [norm_ticker(v["activo"]) for v in inversiones if (v["tipo"] or "").lower() == "cripto"]
    cg_prices = fetch_crypto_prices(crypto_tks)
    live_ar = fetch_live_ar_prices()
    total = 0.0
    for v in inversiones:
        tk = norm_ticker(v["activo"])
        crypto = (v["tipo"] or "").lower() == "cripto"
        precio = v["precioActual"]
        if crypto and cg_prices.get(tk):
            precio = cg_prices[tk]
        elif not crypto and live_ar.get(tk):
            precio = live_ar[tk]
        val = v["cantidad"] * precio
        total += val * blue_rate if crypto else val
    return total


def upsert_historial(valor_ars):
    hist = []
    if HIST_PATH.exists():
        try:
            hist = json.loads(HIST_PATH.read_text(encoding="utf-8"))
        except Exception:
            hist = []
    today = date.today().strftime("%Y-%m-%d")
    idx = next((i for i, h in enumerate(hist) if h.get("fecha") == today), None)
    if idx is not None:
        hist[idx]["valor"] = valor_ars
    else:
        hist.append({"fecha": today, "valor": valor_ars})
    hist.sort(key=lambda h: h["fecha"])
    HIST_PATH.write_text(json.dumps(hist, ensure_ascii=False, indent=None), encoding="utf-8")
    return hist


def main():
    try:
        inversiones, blue_rate = read_inversiones_y_blue()
    except Exception as e:
        log(f"No se pudo leer el Excel hoy: {e}. Sin foto para hoy.")
        return
    blue_rate = fetch_blue_rate(blue_rate)
    valor = compute_portfolio_value_ars(inversiones, blue_rate)
    if not valor:
        log("Valor de portafolio calculado en 0 — no se guarda (probable falla de datos).")
        return
    hist = upsert_historial(valor)
    log(f"Foto de hoy guardada: ${valor:,.0f} ARS ({len(hist)} días en el historial).")

    try:
        subprocess.check_call([sys.executable, str(BASE / "generar_datos_html.py")], cwd=str(BASE))
    except Exception as e:
        log(f"generar_datos_html.py falló: {e}")
        return

    try:
        subprocess.run(["git", "add", "Mi Dashboard de Finanzas.html", "docs/index.html",
                         "Programas/historial_portafolio.json"], cwd=str(RAIZ), check=True)
        commit = subprocess.run(["git", "commit", "-m", "Foto diaria del portafolio"],
                                 cwd=str(RAIZ), capture_output=True, text=True)
        if commit.returncode != 0:
            log("Sin cambios nuevos para subir hoy.")
            return
        subprocess.run(["git", "push"], cwd=str(RAIZ), check=True)
        log("Subido a GitHub OK.")
    except Exception as e:
        log(f"No se pudo subir a GitHub: {e}")


if __name__ == "__main__":
    main()
