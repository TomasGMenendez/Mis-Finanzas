"""
MercadoPago → Excel Finanzas Toto
Uso:  python mercadopago_processor.py account_statement-XXXX.csv
Lee CSV de MP, aplica reglas de mis_reglas.json + config_toto.json, y carga
al Excel Finanzas_Toto.xlsx en las hojas Gastos e Ingresos.
"""
import sys, json, os, re, subprocess
from pathlib import Path

BASE = Path(__file__).parent
RAIZ = BASE.parent  # donde vive Finanzas_Toto.xlsx

# Auto-install deps
def ensure(pkg):
    try: __import__(pkg)
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", pkg, "--quiet"])
ensure("pandas")
ensure("openpyxl")

# Las reglas y la clasificación viven en mercadopago_comun.py, compartidas con
# el CRM (crm_mercadopago.py), para que los dos caminos de carga clasifiquen
# exactamente igual.
import mercadopago_comun as MPC

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import date, datetime

if len(sys.argv) < 2:
    print("Uso: python mercadopago_processor.py <archivo_csv>")
    sys.exit(1)

csv_file = sys.argv[1]
if not os.path.exists(csv_file):
    print(f"❌ No encuentro: {csv_file}")
    sys.exit(1)

# Cargar reglas + config
reglas, privadas = MPC.cargar_reglas()
with open(BASE / "config_toto.json", encoding="utf-8") as f:
    config = json.load(f)

# Cargar CSV
df = MPC.leer_csv(csv_file)

def normalize_sort_value(value):
    if isinstance(value, datetime):
        return value
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    if hasattr(value, "to_pydatetime"):
        return value.to_pydatetime()
    if isinstance(value, (date, datetime)):
        return datetime.combine(value, datetime.min.time())
    return value


df[["tipo","fuente","categoria","nota"]] = df.apply(
    lambda r: pd.Series(MPC.clasificar(r["desc"], r["amt"], reglas)), axis=1)

# Reporte
ingresos = df[df["tipo"]=="INGRESO"]
gastos = df[df["tipo"]=="GASTO"]
revisar = df[df["tipo"]=="REVISAR"]
excluidos = df[df["tipo"].isna()]

print("="*70)
print(f"MP → EXCEL — Archivo: {os.path.basename(csv_file)}")
print("="*70)
print(f"Ingresos:  {len(ingresos):>3} filas | ${ingresos['amt'].sum():>15,.2f} ARS  (NO se cargan — ver nota abajo)")
print(f"Gastos:    {len(gastos):>3} filas | ${gastos['amt'].sum():>15,.2f} ARS")
print(f"REVISAR:   {len(revisar):>3} filas | ${revisar['amt'].sum():>15,.2f} ARS")
print(f"Excluidos: {len(excluidos):>3} filas")
print("\nNOTA: por pedido de Toto, este script NUNCA toca la hoja Ingresos.")
print("      Los ingresos (rendimientos MP, transferencias recibidas) se muestran")
print("      arriba solo de referencia, pero no se escriben al Excel.")

if len(revisar) > 0:
    print(f"\n⚠️  {len(revisar)} transferencia(s) a persona(s) nueva(s) — te voy a preguntar la categoría de cada una al cargarlas.")

# Transferencias a una persona/comercio que ninguna regla reconoce: en vez de
# dejarlas con categoría "?" para que Toto las revise después a mano en el
# Excel, se pregunta la categoría en el momento. La respuesta se guarda como
# regla nueva en mis_reglas_privado.json (nunca en el público, por los
# nombres) para que la próxima vez que aparezca la misma persona ya se
# categorice sola. "categorias_sesion" evita preguntar dos veces por la misma
# persona si aparece más de una vez en el mismo archivo CSV.
categorias_sesion = {}

def preguntar_categoria(desc, monto, fecha):
    nombre = MPC.normalizar_nombre(desc)
    if nombre in categorias_sesion:
        return categorias_sesion[nombre]
    print(f"\n⚠️  Transferencia nueva: \"{desc}\" — ${monto:,.2f} — {fecha.strftime('%d/%m/%Y')}")
    cat = input("   ¿A qué categoría la cargo? (Enter para dejarla \"?\" y revisarla después en Excel) > ").strip()
    if not cat:
        cat = "?"
    else:
        MPC.guardar_regla_privada(privadas, desc, cat)
        print(f"   Guardado — la próxima vez \"{desc}\" se va a cargar sola en \"{cat}\".")
    categorias_sesion[nombre] = cat
    return cat

# Cargar al Excel
xlsx_path = RAIZ / "Finanzas_Toto.xlsx"
wb = load_workbook(xlsx_path)
BLUE = "7CA8FF"; INK = "F5F5F5"; FN = "Segoe UI"

# INGRESOS: deliberadamente NO se tocan. Toto pidió que este script jamás
# escriba en la hoja Ingresos — solo procesa Gastos. No borrar este comentario
# ni volver a agregar el bloque de escritura de Ingresos sin que él lo pida.

# GASTOS (incluye REVISAR)
gas = wb["Gastos"]
MAX_ROW = 504  # mismo límite que usa la fórmula de TOTAL (=SUM(J5:J504))

# Última fila real con datos: hay que escanear TODA la columna, no parar en
# el primer hueco. Un bug viejo paraba en el primer blanco, y eso hacía que
# cargas nuevas se metieran en huecos en el medio de la hoja en vez de ir al
# final — así quedó un quilombo de datos viejos superpuestos (corregido a
# mano el 19/07/2026, ver mis_reglas.json y memoria del proyecto).
last_row = 4
for r in range(5, MAX_ROW + 1):
    if gas.cell(r, 2).value is not None:
        last_row = r
row = last_row + 1

# Anti-duplicados: pedido explícito de Toto — no cargar un movimiento si ya
# existe otro de la misma fecha y misma persona/descripcion en la hoja, sin
# importar si el monto cambió. Repetir el mismo gasto a la misma persona en
# días distintos es válido y debe cargarse igual.
existentes = set()
# Además del dedupe por fecha+descripcion (arriba), se guarda fecha+monto
# de cada fila ya cargada. Sirve para avisar (sin bloquear) cuando una
# transferencia nueva coincide en fecha y monto con una ya cargada pero con
# otra descripcion — el caso real que paso con "Bull Market": la misma
# transferencia quedo cargada dos veces con textos distintos y no se detecto
# hasta que Toto lo noto a mano en el Dashboard. La columna "Posible
# duplicado" del Excel (formula en Gastos, col L) cubre también las cargas
# manuales directas en la hoja, que este script nunca ve.
existentes_monto = {}
for r in range(5, last_row + 1):
    fecha_v = gas.cell(r, 2).value
    desc_v = gas.cell(r, 5).value
    monto_v = gas.cell(r, 9).value
    if fecha_v is not None:
        existentes.add((str(fecha_v)[:10], (desc_v or '').strip().lower()))
        existentes_monto.setdefault((str(fecha_v)[:10], monto_v), []).append(desc_v or '')

agregados = 0
saltados = 0
avisos_monto = []
for _, r_ in pd.concat([gastos, revisar]).iterrows():
    key = (str(r_["fecha"].date()), r_["fuente"][:100].strip().lower())
    if MPC.es_duplicado(existentes, r_["fecha"].date(), r_["fuente"][:100]):
        saltados += 1
        continue
    mkey = (str(r_["fecha"].date()), abs(r_["amt"]))
    if mkey in existentes_monto:
        avisos_monto.append((r_["fecha"].date(), abs(r_["amt"]), r_["fuente"][:100], existentes_monto[mkey]))
    categoria = r_["categoria"]
    if r_["tipo"] == "REVISAR":
        categoria = preguntar_categoria(r_["fuente"][:100], abs(r_["amt"]), r_["fecha"])
    gas.cell(row, 2, r_["fecha"].date()).number_format = 'dd/mm/yyyy'
    gas.cell(row, 4, categoria)
    gas.cell(row, 5, r_["fuente"][:100])
    gas.cell(row, 6, "MercadoPago")
    gas.cell(row, 7, "Variable")
    gas.cell(row, 8, "ARS")
    gas.cell(row, 9, float(abs(r_["amt"])))
    existentes.add(key)
    row += 1
    agregados += 1

# Reordenar por fecha + regenerar fórmulas de Mes (col C) y ARS (col J) para
# TODA la hoja — pedido explícito de Toto (19/07/2026): las filas tienen que
# quedar siempre en orden cronológico, más vieja arriba / más nueva abajo, y
# esto tiene que quedar fijo en cada carga futura, no solo en esta.
last_row2 = 4
for r in range(5, MAX_ROW + 1):
    if gas.cell(r, 2).value is not None:
        last_row2 = r

filas_todas = []
for r in range(5, last_row2 + 1):
    fecha_v = gas.cell(r, 2).value
    if fecha_v is None:
        continue
    filas_todas.append({
        'fecha': fecha_v,
        'categoria': gas.cell(r, 4).value,
        'descripcion': gas.cell(r, 5).value,
        'medio': gas.cell(r, 6).value,
        'tipo': gas.cell(r, 7).value,
        'moneda': gas.cell(r, 8).value,
        'monto': gas.cell(r, 9).value,
    })
filas_todas.sort(key=lambda f: normalize_sort_value(f['fecha']))

for r in range(5, last_row2 + 1):
    for c in range(2, 11):
        gas.cell(r, c).value = None

for i, f in enumerate(filas_todas):
    r = 5 + i
    gas.cell(r, 2, f['fecha']).number_format = 'dd/mm/yyyy'
    gas.cell(r, 3, '=IF(B{0}="","",TEXT(B{0},"mmm-yy"))'.format(r))
    gas.cell(r, 4, f['categoria']).font = Font(name=FN, color=INK, size=10)
    gas.cell(r, 5, f['descripcion']).font = Font(name=FN, color=INK, size=10)
    gas.cell(r, 6, f['medio']).font = Font(name=FN, color=INK, size=10)
    gas.cell(r, 7, f['tipo']).font = Font(name=FN, color=INK, size=10)
    gas.cell(r, 8, f['moneda']).font = Font(name=FN, color=INK, size=10)
    gas.cell(r, 9, f['monto']).font = Font(name=FN, color=BLUE)
    gas.cell(r, 10, '=IF(OR(I{0}="",H{0}=""),"",IF(H{0}="ARS",I{0},I{0}*USDT_ARS))'.format(r))

wb.save(xlsx_path)
print(f"\n✅ Cargado al Excel: {xlsx_path}")
print(f"   {agregados} filas nuevas · {saltados} saltadas por ser duplicado (misma fecha + persona/descripcion)")
print(f"   Hoja reordenada por fecha (vieja→nueva), {len(filas_todas)} filas totales.")

if avisos_monto:
    print("\n⚠️  Posibles duplicados (misma fecha y monto, distinta descripcion) — revisalos a mano:")
    for fecha_a, monto_a, desc_nueva, descs_existentes in avisos_monto:
        print(f"  - {fecha_a} | ${monto_a:>10,.2f} | nueva: \"{desc_nueva}\" · ya existia: \"{descs_existentes[0]}\"")
    print("  (también quedan marcados en la hoja Gastos, columna 'Posible duplicado')")
